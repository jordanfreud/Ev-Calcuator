"""
EV Calculator Performance Dashboard Generator.

Reads bet_log.jsonl and produces a professional Excel workbook with:
  - Overview sheet with KPIs, CLV health, and navigation
  - Pick Log sheet with every graded pick
  - Edge Analysis sheet with ROI by edge bucket and model rank
  - Charts for ROI trend and edge distribution

Uses the excel-generator skill design patterns for professional output.
"""

import json
import math
import os
from datetime import datetime, timezone
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import LOG_PATH, REPORT_TIMEZONE


# ─── Theme: Corporate Blue (finance/betting context) ─────────────────────────
THEME = {
    "primary": "1F4E79",
    "light": "D6E3F0",
    "accent": "1F4E79",
    "chart_colors": ["1F4E79", "2E75B6", "5B9BD5", "9DC3E6", "BDD7EE", "DEEBF7"],
}
SERIF_FONT = "Georgia"
SANS_FONT = "Calibri"

POSITIVE_COLOR = "2E7D32"
NEGATIVE_COLOR = "C62828"
WARNING_COLOR = "F57C00"

# Border definitions
OUTER_BORDER = Side(style="thin", color="D1D1D1")
HEADER_BOTTOM = Side(style="medium", color=THEME["primary"])
INNER_H = Side(style="thin", color="D1D1D1")
NO_BORDER = Side(style=None)


def _apply_data_block_borders(ws, start_row, end_row, start_col, end_col, has_header=True):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            left = OUTER_BORDER if col == start_col else NO_BORDER
            right = OUTER_BORDER if col == end_col else NO_BORDER
            top = OUTER_BORDER if row == start_row else INNER_H
            if has_header and row == start_row:
                bottom = HEADER_BOTTOM
            elif row == end_row:
                bottom = OUTER_BORDER
            else:
                bottom = INNER_H
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _header_style():
    return {
        "font": Font(name=SERIF_FONT, size=10, bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color=THEME["primary"], end_color=THEME["primary"], fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def _data_font():
    return Font(name=SANS_FONT, size=11)


def _read_log() -> List[dict]:
    if not os.path.exists(LOG_PATH):
        return []
    rows = []
    with open(LOG_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


def _american_to_prob(odds: int) -> float:
    if odds > 0:
        return 100.0 / (odds + 100.0)
    return -odds / (-odds + 100.0)


def _bucket_edge(prob_edge) -> str:
    if prob_edge is None:
        return "unknown"
    try:
        pe = float(prob_edge)
    except (TypeError, ValueError):
        return "unknown"
    if pe < 0.01:
        return "<1pp"
    if pe < 0.02:
        return "1-2pp"
    if pe < 0.04:
        return "2-4pp"
    return ">=4pp"


def _bucket_model_rank(rank) -> str:
    if rank is None:
        return "none"
    try:
        r = int(rank)
    except (TypeError, ValueError):
        return "none"
    if r <= 3:
        return "1-3"
    if r <= 6:
        return "4-6"
    if r <= 10:
        return "7-10"
    return "11+"


def generate_dashboard(output_path: str = "ev_dashboard.xlsx") -> str:
    rows = _read_log()
    picks = [r for r in rows if r.get("record_type") in {"pick", "provisional_pick"}]
    graded = [r for r in picks if r.get("graded") and r.get("result") in {"win", "loss", "push"}]

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # OVERVIEW SHEET
    # ═══════════════════════════════════════════════════════════════════════════
    ws_overview = wb.active
    ws_overview.title = "Overview"
    ws_overview.sheet_view.showGridLines = False
    ws_overview.column_dimensions["A"].width = 3

    # Title
    ws_overview.merge_cells("B2:G2")
    ws_overview["B2"] = "EV Calculator Performance Dashboard"
    ws_overview["B2"].font = Font(name=SERIF_FONT, size=20, bold=True, color=THEME["primary"])
    ws_overview.row_dimensions[2].height = 35

    # Subtitle
    ws_overview.merge_cells("B3:G3")
    now_str = datetime.now(REPORT_TIMEZONE).strftime("%B %d, %Y %I:%M %p %Z")
    ws_overview["B3"] = f"Generated: {now_str}"
    ws_overview["B3"].font = Font(name=SANS_FONT, size=10, italic=True, color="666666")

    # ── KPI Section ──
    ws_overview.merge_cells("B5:G5")
    ws_overview["B5"] = "KEY PERFORMANCE INDICATORS"
    ws_overview["B5"].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME["primary"])
    ws_overview["B5"].fill = PatternFill(start_color=THEME["light"], end_color=THEME["light"], fill_type="solid")
    ws_overview.row_dimensions[5].height = 25

    total_picks = len(picks)
    total_graded = len(graded)
    wins = sum(1 for r in graded if r.get("result") == "win")
    losses = sum(1 for r in graded if r.get("result") == "loss")
    pushes = sum(1 for r in graded if r.get("result") == "push")
    pending = total_picks - total_graded
    total_units = sum(float(r.get("units") or 0.0) for r in graded)
    graded_non_push = wins + losses
    win_rate = (wins / graded_non_push * 100) if graded_non_push else 0.0
    roi = (total_units / total_graded * 100) if total_graded else 0.0
    avg_ev = (sum(float(r.get("ev") or 0.0) for r in picks) / total_picks * 100) if total_picks else 0.0

    clv_rows = [r for r in picks if r.get("clv") is not None]
    avg_clv = (sum(float(r.get("clv") or 0.0) for r in clv_rows) / len(clv_rows) * 100) if clv_rows else None

    kpi_headers = ["Metric", "Value"]
    kpis = [
        ("Total Picks", total_picks),
        ("Graded", total_graded),
        ("Pending", pending),
        ("Wins", wins),
        ("Losses", losses),
        ("Win Rate", f"{win_rate:.1f}%"),
        ("Total Units", f"{total_units:+.2f}"),
        ("ROI", f"{roi:+.1f}%"),
        ("Avg EV", f"{avg_ev:.2f}%"),
        ("Avg CLV", f"{avg_clv:.4f}%" if avg_clv is not None else "n/a"),
    ]

    kpi_start_row = 7
    for col_idx, header in enumerate(kpi_headers, start=2):
        cell = ws_overview.cell(row=kpi_start_row, column=col_idx, value=header)
        for attr, val in _header_style().items():
            setattr(cell, attr, val)
    ws_overview.row_dimensions[kpi_start_row].height = 30

    for i, (metric, value) in enumerate(kpis, start=1):
        row = kpi_start_row + i
        ws_overview.cell(row=row, column=2, value=metric).font = Font(name=SANS_FONT, size=11, bold=True)
        val_cell = ws_overview.cell(row=row, column=3, value=value)
        val_cell.font = _data_font()
        val_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws_overview.row_dimensions[row].height = 18

    kpi_end_row = kpi_start_row + len(kpis)
    _apply_data_block_borders(ws_overview, kpi_start_row, kpi_end_row, 2, 3, has_header=True)

    ws_overview.column_dimensions["B"].width = 20
    ws_overview.column_dimensions["C"].width = 18

    # ── CLV Health Check ──
    clv_section_row = kpi_end_row + 3
    ws_overview.merge_cells(f"B{clv_section_row}:G{clv_section_row}")
    ws_overview[f"B{clv_section_row}"] = "CLV HEALTH CHECK"
    ws_overview[f"B{clv_section_row}"].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME["primary"])
    ws_overview[f"B{clv_section_row}"].fill = PatternFill(start_color=THEME["light"], end_color=THEME["light"], fill_type="solid")

    clv_msg_row = clv_section_row + 2
    if avg_clv is not None:
        if avg_clv > 0:
            clv_status = f"HEALTHY: Avg CLV is +{avg_clv:.3f}%. You are consistently beating closing lines."
            clv_color = POSITIVE_COLOR
        elif avg_clv > -0.5:
            clv_status = f"MARGINAL: Avg CLV is {avg_clv:.3f}%. Edge is thin — monitor closely."
            clv_color = WARNING_COLOR
        else:
            clv_status = f"WARNING: Avg CLV is {avg_clv:.3f}%. You are getting worse lines than close. Review timing."
            clv_color = NEGATIVE_COLOR
    else:
        clv_status = "INSUFFICIENT DATA: No CLV data available yet. Run line collection to track."
        clv_color = "666666"

    ws_overview.merge_cells(f"B{clv_msg_row}:G{clv_msg_row}")
    ws_overview[f"B{clv_msg_row}"] = clv_status
    ws_overview[f"B{clv_msg_row}"].font = Font(name=SANS_FONT, size=11, bold=True, color=clv_color)

    # ── Sheet Navigation ──
    nav_row = clv_msg_row + 3
    ws_overview[f"B{nav_row}"] = "CONTENTS"
    ws_overview[f"B{nav_row}"].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME["accent"])

    sheets = [("Overview", "KPIs and CLV health"), ("Pick Log", "Every pick with full detail"), ("Edge Analysis", "ROI by edge bucket and model rank")]
    for i, (name, desc) in enumerate(sheets, start=1):
        r = nav_row + i
        cell = ws_overview.cell(row=r, column=2, value=name)
        cell.hyperlink = f"#'{name}'!A1"
        cell.font = Font(color=THEME["accent"], underline="single")
        ws_overview.cell(row=r, column=3, value=desc).font = Font(name=SANS_FONT, size=10, italic=True, color="666666")

    # ═══════════════════════════════════════════════════════════════════════════
    # PICK LOG SHEET
    # ═══════════════════════════════════════════════════════════════════════════
    ws_log = wb.create_sheet("Pick Log")
    ws_log.sheet_view.showGridLines = False
    ws_log.column_dimensions["A"].width = 3

    ws_log.merge_cells("B2:L2")
    ws_log["B2"] = "Pick Log"
    ws_log["B2"].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME["primary"])
    ws_log.row_dimensions[2].height = 35

    headers = ["Date", "Sport", "Game", "Team", "Book", "Odds", "EV%", "Edge (pp)", "Kelly%", "Result", "Units"]
    header_row = 4
    for col_idx, header in enumerate(headers, start=2):
        cell = ws_log.cell(row=header_row, column=col_idx, value=header)
        for attr, val in _header_style().items():
            setattr(cell, attr, val)
    ws_log.row_dimensions[header_row].height = 30

    sorted_picks = sorted(picks, key=lambda x: x.get("created_at", ""), reverse=True)
    for i, pick in enumerate(sorted_picks):
        row = header_row + 1 + i
        created = pick.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created).astimezone(REPORT_TIMEZONE)
            date_str = dt.strftime("%m/%d/%y")
        except (ValueError, TypeError):
            date_str = created[:10] if len(created) >= 10 else "n/a"

        ev_val = float(pick.get("ev") or 0.0) * 100
        edge_val = float(pick.get("probability_edge") or 0.0) * 100
        kelly_val = float(pick.get("kelly_fraction") or 0.0) * 100
        odds = pick.get("odds")
        odds_str = f"+{odds}" if odds and int(odds) > 0 else str(odds) if odds else "n/a"
        result = pick.get("result") or ("pending" if not pick.get("graded") else "n/a")
        units = pick.get("units")
        units_val = float(units) if units is not None else None

        values = [
            date_str,
            pick.get("sport", ""),
            pick.get("game", ""),
            pick.get("team", ""),
            pick.get("book", ""),
            odds_str,
            ev_val,
            edge_val,
            kelly_val,
            result,
            units_val,
        ]

        for col_idx, val in enumerate(values, start=2):
            cell = ws_log.cell(row=row, column=col_idx, value=val)
            cell.font = _data_font()
            if col_idx in (7, 8, 9):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 12 and val is not None:
                cell.number_format = "+0.00;-0.00;0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    cell.font = Font(name=SANS_FONT, size=11, color=POSITIVE_COLOR if val > 0 else (NEGATIVE_COLOR if val < 0 else "000000"))
            elif col_idx == 11:
                if result == "win":
                    cell.font = Font(name=SANS_FONT, size=11, bold=True, color=POSITIVE_COLOR)
                elif result == "loss":
                    cell.font = Font(name=SANS_FONT, size=11, bold=True, color=NEGATIVE_COLOR)
        ws_log.row_dimensions[row].height = 18

    log_end_row = header_row + len(sorted_picks)
    if sorted_picks:
        _apply_data_block_borders(ws_log, header_row, log_end_row, 2, 12, has_header=True)

    # Column widths
    col_widths = [10, 16, 28, 18, 16, 10, 10, 10, 10, 10, 10]
    for i, w in enumerate(col_widths, start=2):
        ws_log.column_dimensions[get_column_letter(i)].width = w

    if len(sorted_picks) > 10:
        ws_log.freeze_panes = "B5"
    if len(sorted_picks) > 20:
        ws_log.auto_filter.ref = f"B4:{get_column_letter(12)}{log_end_row}"

    # ═══════════════════════════════════════════════════════════════════════════
    # EDGE ANALYSIS SHEET
    # ═══════════════════════════════════════════════════════════════════════════
    ws_edge = wb.create_sheet("Edge Analysis")
    ws_edge.sheet_view.showGridLines = False
    ws_edge.column_dimensions["A"].width = 3

    ws_edge.merge_cells("B2:H2")
    ws_edge["B2"] = "Edge Analysis"
    ws_edge["B2"].font = Font(name=SERIF_FONT, size=18, bold=True, color=THEME["primary"])
    ws_edge.row_dimensions[2].height = 35

    # ── ROI by Edge Bucket ──
    ws_edge.merge_cells("B4:F4")
    ws_edge["B4"] = "ROI BY PROBABILITY EDGE"
    ws_edge["B4"].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME["primary"])
    ws_edge["B4"].fill = PatternFill(start_color=THEME["light"], end_color=THEME["light"], fill_type="solid")
    ws_edge.row_dimensions[4].height = 25

    edge_buckets = {}
    for r in graded:
        bucket = _bucket_edge(r.get("probability_edge"))
        edge_buckets.setdefault(bucket, {"graded": 0, "wins": 0, "units": 0.0})
        edge_buckets[bucket]["graded"] += 1
        if r.get("result") == "win":
            edge_buckets[bucket]["wins"] += 1
        edge_buckets[bucket]["units"] += float(r.get("units") or 0.0)

    edge_headers = ["Edge Bucket", "Graded", "Wins", "Win Rate", "Units", "ROI%"]
    edge_header_row = 6
    for col_idx, header in enumerate(edge_headers, start=2):
        cell = ws_edge.cell(row=edge_header_row, column=col_idx, value=header)
        for attr, val in _header_style().items():
            setattr(cell, attr, val)
    ws_edge.row_dimensions[edge_header_row].height = 30

    bucket_order = ["<1pp", "1-2pp", "2-4pp", ">=4pp", "unknown"]
    edge_data_start = edge_header_row + 1
    row_idx = edge_data_start
    for bucket in bucket_order:
        if bucket not in edge_buckets:
            continue
        data = edge_buckets[bucket]
        g = data["graded"]
        w = data["wins"]
        wr = (w / g * 100) if g else 0
        u = data["units"]
        roi_val = (u / g * 100) if g else 0

        values = [bucket, g, w, f"{wr:.1f}%", round(u, 2), f"{roi_val:+.1f}%"]
        for col_idx, val in enumerate(values, start=2):
            cell = ws_edge.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _data_font()
            if col_idx in (3, 4):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_idx == 7:
                color = POSITIVE_COLOR if roi_val > 0 else (NEGATIVE_COLOR if roi_val < 0 else "000000")
                cell.font = Font(name=SANS_FONT, size=11, bold=True, color=color)
        ws_edge.row_dimensions[row_idx].height = 18
        row_idx += 1

    edge_data_end = row_idx - 1
    if edge_data_end >= edge_data_start:
        _apply_data_block_borders(ws_edge, edge_header_row, edge_data_end, 2, 7, has_header=True)

    # ── ROI by Model Rank ──
    rank_section_row = edge_data_end + 3 if edge_data_end >= edge_data_start else edge_header_row + 3
    ws_edge.merge_cells(f"B{rank_section_row}:F{rank_section_row}")
    ws_edge[f"B{rank_section_row}"] = "ROI BY MODEL RANK"
    ws_edge[f"B{rank_section_row}"].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME["primary"])
    ws_edge[f"B{rank_section_row}"].fill = PatternFill(start_color=THEME["light"], end_color=THEME["light"], fill_type="solid")

    rank_buckets = {}
    for r in graded:
        bucket = _bucket_model_rank(r.get("model_rank"))
        rank_buckets.setdefault(bucket, {"graded": 0, "wins": 0, "units": 0.0})
        rank_buckets[bucket]["graded"] += 1
        if r.get("result") == "win":
            rank_buckets[bucket]["wins"] += 1
        rank_buckets[bucket]["units"] += float(r.get("units") or 0.0)

    rank_header_row = rank_section_row + 2
    for col_idx, header in enumerate(edge_headers, start=2):
        cell = ws_edge.cell(row=rank_header_row, column=col_idx, value=header.replace("Edge Bucket", "Model Rank"))
        for attr, val in _header_style().items():
            setattr(cell, attr, val)
    ws_edge.row_dimensions[rank_header_row].height = 30

    rank_order = ["1-3", "4-6", "7-10", "11+", "none"]
    rank_data_start = rank_header_row + 1
    row_idx = rank_data_start
    for bucket in rank_order:
        if bucket not in rank_buckets:
            continue
        data = rank_buckets[bucket]
        g = data["graded"]
        w = data["wins"]
        wr = (w / g * 100) if g else 0
        u = data["units"]
        roi_val = (u / g * 100) if g else 0

        values = [bucket, g, w, f"{wr:.1f}%", round(u, 2), f"{roi_val:+.1f}%"]
        for col_idx, val in enumerate(values, start=2):
            cell = ws_edge.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _data_font()
            if col_idx == 7:
                color = POSITIVE_COLOR if roi_val > 0 else (NEGATIVE_COLOR if roi_val < 0 else "000000")
                cell.font = Font(name=SANS_FONT, size=11, bold=True, color=color)
        ws_edge.row_dimensions[row_idx].height = 18
        row_idx += 1

    rank_data_end = row_idx - 1
    if rank_data_end >= rank_data_start:
        _apply_data_block_borders(ws_edge, rank_header_row, rank_data_end, 2, 7, has_header=True)

    # ── KEY INSIGHTS ──
    insights_row = rank_data_end + 3 if rank_data_end >= rank_data_start else rank_header_row + 3
    ws_edge.merge_cells(f"B{insights_row}:G{insights_row}")
    ws_edge[f"B{insights_row}"] = "KEY INSIGHTS"
    ws_edge[f"B{insights_row}"].font = Font(name=SERIF_FONT, size=14, bold=True, color=THEME["primary"])

    insights = []
    if total_graded > 0:
        insights.append(f"Record: {wins}W-{losses}L ({win_rate:.1f}% win rate) across {total_graded} graded picks.")
        insights.append(f"Total units: {total_units:+.2f} | ROI: {roi:+.1f}%")
    if avg_clv is not None:
        if avg_clv > 0:
            insights.append(f"CLV is positive (+{avg_clv:.3f}%) — you are beating closing lines consistently.")
        else:
            insights.append(f"CLV is negative ({avg_clv:.3f}%) — consider scanning earlier when lines are softer.")
    if not insights:
        insights.append("No graded picks yet. Run the scanner and check back after games complete.")

    for i, insight in enumerate(insights):
        r = insights_row + 1 + i
        ws_edge.merge_cells(f"B{r}:G{r}")
        ws_edge[f"B{r}"] = insight
        ws_edge[f"B{r}"].font = Font(name=SANS_FONT, size=11, color="333333")

    # Column widths for Edge Analysis
    edge_col_widths = [14, 10, 10, 12, 12, 12]
    for i, w in enumerate(edge_col_widths, start=2):
        ws_edge.column_dimensions[get_column_letter(i)].width = w

    # ── Bar Chart: ROI by Edge Bucket ──
    if edge_data_end >= edge_data_start:
        chart = BarChart()
        chart.title = "ROI% by Probability Edge Bucket"
        chart.style = 10
        chart.y_axis.title = "ROI %"
        chart.x_axis.title = "Edge Bucket"
        chart.width = 18
        chart.height = 12

        # Use units column for chart data
        data_ref = Reference(ws_edge, min_col=6, min_row=edge_header_row, max_row=edge_data_end)
        cats_ref = Reference(ws_edge, min_col=2, min_row=edge_data_start, max_row=edge_data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        if chart.series:
            chart.series[0].graphicalProperties.solidFill = THEME["chart_colors"][0]

        chart_row = insights_row + len(insights) + 3
        ws_edge.add_chart(chart, f"B{chart_row}")

    # Save
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = generate_dashboard()
    print(f"Dashboard saved to: {path}")
